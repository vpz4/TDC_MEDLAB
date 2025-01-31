# -*- coding: utf-8 -*-
"""
Created on Wed Sep  5 17:17:59 2018

@author: vpezoulas
"""

import Orange
import numpy as np
import pandas as pd
import re
import sys
import io
import json
import scipy
import timeit
import itertools
import os
import random
import csv
import openpyxl
import warnings
from io import StringIO, BytesIO
from datetime import datetime
from scipy.stats import spearmanr, pearsonr, kendalltau
from flask import Flask, jsonify, request
from io import StringIO
from outliers import smirnov_grubbs as grubbs
from Orange.preprocess import Impute, Average
from collections import Counter
from Levenshtein import jaro
from xlrd import open_workbook
from flask import render_template
from xlutils.copy import copy
from openpyxl import Workbook
from sklearn.ensemble import IsolationForest
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
warnings.filterwarnings('ignore')

#application specs
HOST_FOLDER = os.path.dirname(os.path.abspath(__file__))+'/hosted/'
RESULTS_FOLDER = os.path.dirname(os.path.abspath(__file__))+'/results/'
ALLOWED_EXTENSIONS = set(['xlsx','xls','csv', 'json'])

#application specs
app = Flask(__name__)
app.config['RESULTS_FOLDER'] = RESULTS_FOLDER
app.secret_key = 'some_secret'

#capture print screen
class Capturing(list):
    def __enter__(self):
        self._stdout = sys.stdout
        sys.stdout = self._stringio = StringIO()
        return self
    def __exit__(self, *args):
        self.extend(self._stringio.getvalue().splitlines())
        del self._stringio  # free up some memory
        sys.stdout = self._stdout


#assistance for json interlinking
def mangle(s):
    return s.strip()[1:-1]


#connect json files    
def cat_json(output_filename, input_filenames):
    with open(output_filename, "w") as outfile:
        first = True
        for infile_name in input_filenames:
            with open(infile_name) as infile:
                if first:
                    outfile.write('[')
                    first = False
                else:
                    outfile.write(',')
                outfile.write(mangle(infile.read()))
        outfile.write(']')
                        
                        
def formatNumber(num):
  if num % 1 == 0:
    return int(num)
  else:
    return num

  
def formatNumber_v2(num):
  if num % 1 == 0:
    return [int(num), 1]
  else:
    return [num, 0]


def formatNumber_v3(num):
    try:
        y = float(num)
        if(y % 1 == 0):
            return int(y)
        else:
            return y
    except:
        return num


def intersect(seq1, seq2):
    res = []                     # start empty
    for x in seq1:               # scan seq1
        if x in seq2:            # common item?
            res.append(x)        # add to end
    return res


def create_wr_io(path_name, pythonDictionary):
    with io.open(path_name, 'w', encoding='utf-8') as f:
        f.write(json.dumps(pythonDictionary, ensure_ascii=True, sort_keys=False, indent=4))
    
    with open(path_name) as json_data:
        d = json.load(json_data)
    
    return d


def outliers_iqr(ys):
    [quartile_1, quartile_3] = np.percentile(ys, [25, 75])
    iqr = quartile_3 - quartile_1
    lower_bound = quartile_1 - (iqr * 1.5)
    upper_bound = quartile_3 + (iqr * 1.5)
    outliers_ind = np.where((ys > upper_bound) | (ys < lower_bound))
    return [iqr, outliers_ind]


def outliers_z_score(ys):
    threshold = 3
    mean_y = np.mean(ys)
    stdev_y = np.std(ys)
    z_scores = [(y - mean_y) / stdev_y for y in ys]
    outliers_ind = np.where(np.abs(z_scores) > threshold)
    return [z_scores, outliers_ind]


def outliers_z_score_mod(ys):
    threshold = 3.5
    median_y = np.median(ys)
    MAD = np.median(np.abs(ys-np.median(ys)))
    meanAD = np.mean(np.abs(ys-np.mean(ys)))
    
    if(MAD == 0):
        z_scores = [(y - median_y) / (1.253314*meanAD) for y in ys]
    else:
        z_scores = [(y - median_y) / (1.486*MAD) for y in ys]
       
    outliers_ind = np.where(np.abs(z_scores) > threshold)
    return [z_scores, outliers_ind]


def outliers_mad(ys):
    mad = np.median(np.abs(ys-np.median(ys)))
    mad_scores = [(y - mad) for y in ys]
    outliers_ind = np.where(np.abs(mad_scores) > 3*mad)
    return [mad_scores, outliers_ind]


def write_evaluation_report(data_org, r, c, features_total, metas_features, pos_metas, ranges_r, var_type_final, var_type_final_2, 
                            var_type_metas, var_type_metas_2, features_state_metas, incompatibilities_metas, features_missing_values_metas, 
                            bad_features_metas, bad_features_ind_metas,fair_features_metas, fair_features_ind_metas, 
                            good_features_metas, good_features_ind_metas, a_total_metas, outliers_ind_metas, y_score_metas, outliers_pos,
                            outliers_pos_metas, ranges_metas_r, features_missing_values, features_state, outliers_ind, incompatibilities, a_total, path_f):
    
    ranges = []
    for i in range(len(ranges_r)):
        if(len(ranges_r[i]) < 100):
            ranges.append(ranges_r[i])
        else:
            ranges.append('Too large to display!!')

    ranges_metas = []
    for i in range(len(ranges_metas_r)):
        if(len(ranges_metas_r[i]) < 100):
            ranges_metas.append(ranges_metas_r[i])
        else:
            ranges_metas.append('Too large to display!!')
    
    count_i = 0
    for i in range(0,len(outliers_pos)):
        if(outliers_pos[i] != '-'):
            count_i += len(outliers_pos[i])
            
    count_ii = 0
    for i in range(0,len(outliers_pos_metas)):
        if(outliers_pos_metas[i] != '-'):
            count_ii += len(outliers_pos_metas[i])        
    
    wb = Workbook()
    sheet1 = wb.create_sheet('Evaluation',0)
    # sheet_names = wb.sheetnames
    # sheet1 = sheet_names['Sheet1']
    
    # print(["Metas pos:", pos_metas])
    # print("")
    
    offs_1 = 1 #<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
    offs_2 = 1 #<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
    
    #panel_1
    n_disc = Counter(var_type_final_2+var_type_metas_2)['categorical']
    n_cont = Counter(var_type_final_2+var_type_metas_2)['numeric']
    n_unkn = Counter(var_type_final_2+var_type_metas_2)['unknown']
    n_miss = np.around(100*((a_total+a_total_metas)/((c+len(metas_features))*r)), 2)

    my_gray = openpyxl.styles.colors.Color(rgb='00D9D9D9')
    my_aristo = openpyxl.styles.colors.Color(rgb='00CCC0DA')
    my_bad = openpyxl.styles.colors.Color(rgb='00FF99CC')
    my_fair = openpyxl.styles.colors.Color(rgb='00CCFFCC')
    my_good = openpyxl.styles.colors.Color(rgb='00CCCCFF')
    my_inco = openpyxl.styles.colors.Color(rgb='00FF99CC')
    my_no_inco = openpyxl.styles.colors.Color(rgb='00CCCCFF')
    my_yes = openpyxl.styles.colors.Color(rgb='00FF99CC')
    my_no = openpyxl.styles.colors.Color(rgb='00CCCCFF')
    my_non_applicable = openpyxl.styles.colors.Color(rgb='00D9D9D9')
    
    style0 = PatternFill(patternType='solid', fgColor=my_gray)
    my_font0 = Font(size=14, bold=True, name='Calibri')
    
    my_font1_1 = Font(size=11, bold=True, name='Calibri')
    
    style1_2 = PatternFill(patternType='solid', fgColor=my_aristo) #aristo
    
    my_font2 = Font(size=11, name='Calibri')
    style2_2 = PatternFill(patternType='solid', fgColor=my_bad) #bad
    style2_3 = PatternFill(patternType='solid', fgColor=my_fair) #fair
    style2_4 = PatternFill(patternType='solid', fgColor=my_good) #good

    my_font3 = Font(size=11, name='Calibri')
    style3_2 = PatternFill(patternType='solid', fgColor=my_yes) #yes
    style3_3 = PatternFill(patternType='solid', fgColor=my_no) #no
    style3_4 = PatternFill(patternType='solid', fgColor=my_non_applicable) #not-applicable

    my_font4 = Font(size=11, name='Calibri')
    style4_2 = PatternFill(patternType='solid', fgColor=my_inco) #yes inco
    style4_3 = PatternFill(patternType='solid', fgColor=my_no_inco) #no inco
    
    sheet1.cell(offs_1,offs_2,"Metadata")
    sheet1.cell(offs_1,offs_2).fill = style0
    sheet1.cell(offs_1,offs_2).font = my_font0
    
    sheet1.cell(offs_1+1,offs_2,"Number of feature(s)")
    sheet1.cell(offs_1+1,offs_2).fill = style1_2
    sheet1.cell(offs_1+1,offs_2).font = my_font1_1
    sheet1.cell(offs_1+1,offs_2+1,str(c+len(metas_features)))
    
    sheet1.cell(offs_1+2,offs_2,"Number of instance(s)")
    sheet1.cell(offs_1+2,offs_2).fill = style1_2
    sheet1.cell(offs_1+2,offs_2).font = my_font1_1
    sheet1.cell(offs_1+2,offs_2+1,str(r))
    
    sheet1.cell(offs_1+3,offs_2,"Discrete feature(s)")
    sheet1.cell(offs_1+3,offs_2).fill = style1_2
    sheet1.cell(offs_1+3,offs_2).font = my_font1_1     
    sheet1.cell(offs_1+3,offs_2+1,str(n_disc))   
    
    sheet1.cell(offs_1+4,offs_2,"Continuous feature(s)")
    sheet1.cell(offs_1+4,offs_2).fill = style1_2
    sheet1.cell(offs_1+4,offs_2).font = my_font1_1
    sheet1.cell(offs_1+4,offs_2+1,str(n_cont))    
    
    sheet1.cell(offs_1+5,offs_2,"Unknown feature(s)")
    sheet1.cell(offs_1+5,offs_2).fill = style1_2
    sheet1.cell(offs_1+5,offs_2).font = my_font1_1      
    sheet1.cell(offs_1+5,offs_2+1,str(n_unkn))   
    
    sheet1.cell(offs_1+6,offs_2,"Missing values (%)")
    sheet1.cell(offs_1+6,offs_2).fill = style1_2
    sheet1.cell(offs_1+6,offs_2).font = my_font1_1
    sheet1.cell(offs_1+6,offs_2+1,str(n_miss)+"%")
    
    #panel_2
    sheet1.cell(offs_1+11,offs_2,"Quality assessment")
    sheet1.cell(offs_1+11,offs_2).fill = style0
    sheet1.cell(offs_1+11,offs_2).font = my_font0
    c1=0
    sheet1.cell(c1+13,offs_2,"Features").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2).fill = style1_2
    sheet1.cell(c1+13,offs_2).font = my_font1_1
    for i in range(c+len(metas_features)):
        sheet1.cell(c1+14,offs_2,features_total[i])     
        c1=c1+1
    
    c1=0 
    c2=0
    c3=0
    sheet1.cell(c1+13,offs_2+1,"Value range").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2+1).fill = style1_2
    sheet1.cell(c1+13,offs_2+1).font = my_font1_1
    
    # print("Total length of ranges = ", len(ranges))
    # print("Total length of ranges metas =", len(ranges_metas))
    # print("Total number of features =", len(features_total))
    # print("Total number of features metas =", len(metas_features))
    # print("Length of pos metas =", len(pos_metas))
    # print("Total features =", features_total)
    # print("Metas features =", metas_features)
    
    # print("Ranges")
    # print(ranges)
    
    # print("Ranges metas")
    # print(ranges_metas)
    
    for i in range(c+len(metas_features)):
        # print("")
        # print("Working for index", str(i))
        # print("Working for feature", str(features_total[i]))
        if(i not in pos_metas):
            # print("Non meta index", str(i), "in pos")
            # print("c2 index", str(c2))
            ranges_f = str(ranges[c2]).replace("'",'')
            # print("Range", str(ranges_f))
            sheet1.cell(c1+14,offs_2+1,ranges_f).alignment = Alignment(horizontal='center')     
            c2 = c2+1
        else:
            # print("Meta index", str(i), "in pos metas")
            ranges_f = str(ranges_metas[c3]).replace("'",'')
            # print("Range", str(ranges_f))
            sheet1.cell(c1+14,offs_2+1,ranges_f).alignment = Alignment(horizontal='center')
            c3 = c3+1
        c1=c1+1
        
    c1=0 
    c2=0 
    c3=0
    sheet1.cell(c1+13,offs_2+2,"Type").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2+2).fill = style1_2
    sheet1.cell(c1+13,offs_2+2).font = my_font1_1
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            sheet1.cell(c1+14,offs_2+2,var_type_final_2[c2]).alignment = Alignment(horizontal='center')           
            c2 = c2+1
        else:
            sheet1.cell(c1+14,offs_2+2,var_type_metas_2[c3]).alignment = Alignment(horizontal='center')
            c3 = c3+1
        c1=c1+1
    
    c1=0 
    c2=0 
    c3=0
    sheet1.cell(c1+13,offs_2+3,"Variable type").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2+3).fill = style1_2
    sheet1.cell(c1+13,offs_2+3).font = my_font1_1
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            sheet1.cell(c1+14,offs_2+3,var_type_final[c2]).alignment = Alignment(horizontal='center')
            c2 = c2+1
        else:
            sheet1.cell(c1+14,offs_2+3,var_type_metas[c3]).alignment = Alignment(horizontal='center')     
            c3 = c3+1
        c1=c1+1
        
    c1=0 
    c2=0 
    c3=0
    sheet1.cell(c1+13,offs_2+4,"Missing values").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2+4).fill = style1_2
    sheet1.cell(c1+13,offs_2+4).font = my_font1_1
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            sheet1.cell(c1+14,offs_2+4,features_missing_values[c2]).alignment = Alignment(horizontal='center')            
            c2 = c2+1
        else:
            sheet1.cell(c1+14,offs_2+4,features_missing_values_metas[c3]).alignment = Alignment(horizontal='center')          
            c3 = c3+1
        c1=c1+1

    c1=0 
    c2=0 
    c3=0
    sheet1.cell(c1+13,offs_2+5,"State").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2+5).fill = style1_2
    sheet1.cell(c1+13,offs_2+5).font = my_font1_1
    fair_counter = 0
    good_counter = 0
    bad_counter = 0
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            if(features_state[c2] == "bad"):
                sheet1.cell(c1+14,offs_2+5,features_state[c2]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+5).fill = style2_2
                sheet1.cell(c1+14,offs_2+5).font = my_font2 
                bad_counter+=1
            elif(features_state[c2] == "fair"):
                sheet1.cell(c1+14,offs_2+5,features_state[c2]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+5).fill = style2_3
                sheet1.cell(c1+14,offs_2+5).font = my_font2
                fair_counter+=1
            elif(features_state[c2] == "good"):
                sheet1.cell(c1+14,offs_2+5,features_state[c2]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+5).fill = style2_4
                sheet1.cell(c1+14,offs_2+5).font = my_font2
                good_counter+=1
            c2=c2+1
        else:
            if(features_state_metas[c3] == "bad"):
                sheet1.cell(c1+14,offs_2+5,features_state_metas[c3]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+5).fill = style2_2
                sheet1.cell(c1+14,offs_2+5).font = my_font2
                bad_counter+=1
            elif(features_state_metas[c3] == "fair"):
                sheet1.cell(c1+14,offs_2+5,features_state_metas[c3]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+5).fill = style2_3
                sheet1.cell(c1+14,offs_2+5).font = my_font2
                fair_counter+=1
            elif(features_state_metas[c3] == "good"):
                sheet1.cell(c1+14,offs_2+5,features_state_metas[c3]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+5).fill = style2_4
                sheet1.cell(c1+14,offs_2+5).font = my_font2
                good_counter+=1
            c3=c3+1
        c1=c1+1
        
    ###########################################################################
    sheet1.cell(offs_1+7,offs_2,"Good feature(s) (%)")
    sheet1.cell(offs_1+7,offs_2).fill = style1_2
    sheet1.cell(offs_1+7,offs_2).font = my_font1_1
    sheet1.cell(offs_1+7,offs_2+1,str(good_counter)+" ("+str(np.around((good_counter/(c+len(metas_features)))*100,1))+"%)")

    sheet1.cell(offs_1+8,offs_2,"Fair feature(s) (%)")
    sheet1.cell(offs_1+8,offs_2).fill = style1_2
    sheet1.cell(offs_1+8,offs_2).font = my_font1_1
    sheet1.cell(offs_1+8,offs_2+1,str(fair_counter)+" ("+str(np.around((fair_counter/(c+len(metas_features)))*100,1))+"%)")

    sheet1.cell(offs_1+9,offs_2,"Bad feature(s) (%)")
    sheet1.cell(offs_1+9,offs_2).fill = style1_2
    sheet1.cell(offs_1+9,offs_2).font = my_font1_1
    sheet1.cell(offs_1+9,offs_2+1,str(bad_counter)+" ("+str(np.around((bad_counter/(c+len(metas_features)))*100,1))+"%)")
    
    sheet1.cell(offs_1+10,offs_2,"Outlier(s) (%)")
    sheet1.cell(offs_1+10,offs_2).fill = style1_2
    sheet1.cell(offs_1+10,offs_2).font = my_font1_1
    sheet1.cell(offs_1+10,offs_2+1,str(count_i+count_ii)+" ("+str(np.around(((count_i+count_ii)/(r*(c+len(metas_features))))*100,2))+"%)")
    ###########################################################################
    
    c1=0 
    c2=0 
    c3=0
    sheet1.cell(c1+13,offs_2+6,"Outliers").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2+6).fill = style1_2
    sheet1.cell(c1+13,offs_2+6).font = my_font1_1
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            if(outliers_ind[c2] == "yes"):
                sheet1.cell(c1+14,offs_2+6,outliers_ind[c2]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+6).fill = style3_2
                sheet1.cell(c1+14,offs_2+6).font = my_font3
            elif(outliers_ind[c2] == "no"):
                sheet1.cell(c1+14,offs_2+6,outliers_ind[c2]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+6).fill = style3_3
                sheet1.cell(c1+14,offs_2+6).font = my_font3
            elif(outliers_ind[c2] == "not-applicable"):
                sheet1.cell(c1+14,offs_2+6,outliers_ind[c2]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+6).fill = style3_4
                sheet1.cell(c1+14,offs_2+6).font = my_font3
            c2=c2+1
        else:
            if(outliers_ind_metas[c3] == "yes"):
                sheet1.cell(c1+14,offs_2+6,outliers_ind_metas[c3]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+6).fill = style3_2
                sheet1.cell(c1+14,offs_2+6).font = my_font3
            elif(outliers_ind_metas[c3] == "no"):
                sheet1.cell(c1+14,offs_2+6,outliers_ind_metas[c3]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+6).fill = style3_3
                sheet1.cell(c1+14,offs_2+6).font = my_font3
            elif(outliers_ind_metas[c3] == "not-applicable"):
                sheet1.cell(c1+14,offs_2+6,outliers_ind_metas[c3]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+6).fill = style3_4
                sheet1.cell(c1+14,offs_2+6).font = my_font3
            c3=c3+1
        c1=c1+1

    c1=0 
    c2=0 
    c3=0
    sheet1.cell(c1+13,offs_2+7,"Incompatibilities").alignment = Alignment(horizontal='center')
    sheet1.cell(c1+13,offs_2+7).fill = style1_2
    sheet1.cell(c1+13,offs_2+7).font = my_font1_1     
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            if(incompatibilities[c2] == "no"):
                if(features_state[c2] == "bad"):
                    sheet1.cell(c1+14,offs_2+7,'yes, bad feature').alignment = Alignment(horizontal='center')
                    sheet1.cell(c1+14,offs_2+7).fill = style4_2
                    sheet1.cell(c1+14,offs_2+7).font = my_font4                    
                else:
                    sheet1.cell(c1+14,offs_2+7,incompatibilities[c2]).alignment = Alignment(horizontal='center')
                    sheet1.cell(c1+14,offs_2+7).fill = style4_3
                    sheet1.cell(c1+14,offs_2+7).font = my_font4                     
            elif(incompatibilities[c2] == "yes, unknown type of data"):
                sheet1.cell(c1+14,offs_2+7,incompatibilities[c2]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+7).fill = style4_2
                sheet1.cell(c1+14,offs_2+7).font = my_font4             
            c2=c2+1
        else:
            if(incompatibilities_metas[c3] == "no"):
                if(features_state_metas[c3] == "bad"):
                    sheet1.cell(c1+14,offs_2+7,'yes, bad feature').alignment = Alignment(horizontal='center')
                    sheet1.cell(c1+14,offs_2+7).fill = style4_2
                    sheet1.cell(c1+14,offs_2+7).font = my_font4                       
                else:
                    sheet1.cell(c1+14,offs_2+7,incompatibilities_metas[c3]).alignment = Alignment(horizontal='center')
                    sheet1.cell(c1+14,offs_2+7).fill = style4_3
                    sheet1.cell(c1+14,offs_2+7).font = my_font4                    
            elif(incompatibilities_metas[c3] == "yes, unknown type of data"):
                sheet1.cell(c1+14,offs_2+7,incompatibilities_metas[c3]).alignment = Alignment(horizontal='center')
                sheet1.cell(c1+14,offs_2+7).fill = style4_2
                sheet1.cell(c1+14,offs_2+7).font = my_font4              
            c3=c3+1            
        c1=c1+1
    
    if not os.path.exists('results'):
        os.makedirs('results')
    
    width_off = 40
    sheet1.column_dimensions['A'].width = width_off+10
    sheet1.column_dimensions['B'].width = width_off*2
    sheet1.column_dimensions['C'].width = width_off/2
    sheet1.column_dimensions['D'].width = width_off/2
    sheet1.column_dimensions['E'].width = width_off/2
    sheet1.column_dimensions['F'].width = width_off/2
    sheet1.column_dimensions['G'].width = width_off/2
    sheet1.column_dimensions['H'].width = (width_off/2)+10

    sheet1.cell(1,2).fill = style0
    sheet1.cell(1,3).fill = style0
    sheet1.cell(1,4).fill = style0
    sheet1.cell(1,5).fill = style0
    sheet1.cell(1,6).fill = style0
    sheet1.cell(1,7).fill = style0
    sheet1.cell(1,8).fill = style0
    
    sheet1.cell(12,2).fill = style0
    sheet1.cell(12,3).fill = style0
    sheet1.cell(12,4).fill = style0
    sheet1.cell(12,5).fill = style0
    sheet1.cell(12,6).fill = style0
    sheet1.cell(12,7).fill = style0
    sheet1.cell(12,8).fill = style0
    
    wb.save(path_f)


def write_evaluation_report_dict(data_org, r, c, features_total, metas_features, pos_metas, ranges, var_type_final, var_type_final_2, 
                                var_type_metas, var_type_metas_2, features_state_metas, incompatibilities_metas, features_missing_values_metas, 
                                bad_features_metas, bad_features_ind_metas, fair_features_metas, fair_features_ind_metas, 
                                good_features_metas, good_features_ind_metas, a_total_metas, outliers_ind_metas, y_score_metas, 
                                outliers_pos_metas, ranges_metas, features_missing_values, features_state, outliers_ind, incompatibilities, 
                                a_total, means, means_metas, outlier_detection_method_id, imputation_method_id):
    
    #panel_1
    n_disc = Counter(var_type_final_2+var_type_metas_2)['categorical']
    n_cont = Counter(var_type_final_2+var_type_metas_2)['numeric']
    n_unkn = Counter(var_type_final_2+var_type_metas_2)['unknown']
    n_miss = np.around(100*((a_total+a_total_metas)/((c+len(metas_features))*r)), 2)
    
    nof = str(c+len(metas_features)) #Number of feature(s), style2_1
    noi = str(r) #Number of instance(s), style2_1
    df = str(n_disc) #Discrete feature(s), style2_1
    cf = str(n_cont) #Continuous feature(s), style2_1
    uf = str(n_unkn) #Unknown feature(s), style2_1
    mv = str(n_miss)+"%"+' ('+str(a_total+a_total_metas)+')' #Missing values (%), style2_1
    
    #panel_2
    features = []
    for i in range(c+len(metas_features)):
        features.append(features_total[i]) #style2_1_0
    
    value_range = []
    c2=0; c3=0
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            ranges_f = str(ranges[c2]).replace("'",''); c2 = c2+1
            value_range.append(ranges_f) #style2_1
        else:
            ranges_f = str(ranges_metas[c3]).replace("'",''); c3 = c3+1
            value_range.append(ranges_f) #style2_1
        
    c2=0; c3=0
    types = []
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            types.append(var_type_final_2[c2]); c2 = c2+1 #style2_1
        else:
            types.append(var_type_metas_2[c3]); c3 = c3+1 #style2_1
        
    c2=0; c3=0
    var_type = []
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            var_type.append(var_type_final[c2]); c2 = c2+1 #style2_1
        else:
            var_type.append(var_type_metas[c3]); c3 = c3+1 #style2_1
        
    c2=0; c3=0
    m_values = []
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            m_values.append(features_missing_values[c2]); c2 = c2+1 #style2_1
        else:
            m_values.append(features_missing_values_metas[c3]); c3 = c3+1 #style2_1
        
    c2=0; c3=0
    mmf = []
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            mmf.append(means[c2]); c2 = c2+1 #style2_1
        else:
            mmf.append(means_metas[c3]); c3 = c3+1 #style2_1
        
    c2=0; c3=0
    state = []
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            if(features_state[c2] == "bad"):
                state.append(features_state[c2]) #style2_2
            elif(features_state[c2] == "fair"):
                state.append(features_state[c2]) #style2_3
            elif(features_state[c2] == "good"):
                state.append(features_state[c2]) #style2_4
            c2=c2+1
        else:
            if(features_state_metas[c3] == "bad"):
                state.append(features_state_metas[c3]) #style2_2
            elif(features_state_metas[c3] == "fair"):
                state.append(features_state_metas[c3]) #style2_3
            elif(features_state_metas[c3] == "good"):
                state.append(features_state_metas[c3]) #style2_4           
            c3=c3+1
        
    c2=0; c3=0
    out = []
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            if(outliers_ind[c2] == "yes"):
                out.append(outliers_ind[c2]) #style2_2
            elif(outliers_ind[c2] == "no"):
                out.append(outliers_ind[c2]) #style2_3
            elif(outliers_ind[c2] == "not-applicable"):
                out.append(outliers_ind[c2]) #style2_2
            c2=c2+1
        else:
            if(outliers_ind_metas[c3] == "yes"):
                out.append(outliers_ind_metas[c3]) #style2_2
            elif(outliers_ind_metas[c3] == "no"):
                out.append(outliers_ind_metas[c3]) #style2_4
            elif(outliers_ind_metas[c3] == "not-applicable"):
                out.append(outliers_ind_metas[c3]) #style2_2
            c3=c3+1

    c2=0; c3=0
    inco = []
    for i in range(c+len(metas_features)):
        if(i not in pos_metas):
            if(incompatibilities[c2] == "no"):
                if(features_state[c2] == "bad"):
                    inco.append('yes, bad feature') #style2_2
                else:
                    inco.append(incompatibilities[c2]); #style2_4
            elif(incompatibilities[c2] == "yes, unknown type of data"):
                inco.append(incompatibilities[c2]) #style2_2
            c2=c2+1
        else:
            if(incompatibilities_metas[c3] == "no"):
                if(features_state_metas[c3] == "bad"):
                    inco.append('yes, bad feature') #style2_2
                else:
                    inco.append(incompatibilities_metas[c3]); #style2_4
            elif(incompatibilities_metas[c3] == "yes, unknown type of data"):
                inco.append(incompatibilities_metas[c3]) #style2_2
            c3=c3+1

    bf_ind = [i for i,x in enumerate(state) if x=='bad']
    bf_names = [features[i] for i in np.asarray(bf_ind,int)]
    
    python_dict = [{'Number of feature(s)':nof,
                    'Number of instance(s)':noi,
                    'Discrete feature(s)':df,
                    'Continuous feature(s)':cf,
                    'Unknown feature(s)':uf,
                    'Missing values':mv,
                    'Features':features,
                    'Value range':value_range,
                    'Type':types,
                    'Variable type':var_type,
                    'Missing values (per feature)':[str(e) for e in m_values],
                    'Mean/most freq':mmf,
                    'State':state,
                    'Outlier detection method ID':outlier_detection_method_id,
                    'Outliers':out,
                    'Incompatibilities':inco,
                    'Bad features':bf_names,
                    'Imputation method ID':imputation_method_id,
                    'Meta-attribute(s)':metas_features,
                    'Color for good features': 'blue',
                    'Color for fair features': 'green',
                    'Color for bad features': 'red'}]
    return python_dict


def write_curated_dataset(data_org, wb2, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                          var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                          var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, path_f, var_type_final_2):

    wb = Workbook()
    sheet_names = wb.sheetnames
    sheet1 = wb[sheet_names[0]]
    
    # sheet1.set_column(1, np.size(data_org,1), 18)
    
    style01 = openpyxl.styles.colors.Color(rgb='00ECE338') #outlier
    style02 = openpyxl.styles.colors.Color(rgb='00D9D9D9') #imputed
    style03 = openpyxl.styles.colors.Color(rgb='00FF99CC') #bad
    style04 = openpyxl.styles.colors.Color(rgb='00CCFFCC') #fair
    style05 = openpyxl.styles.colors.Color(rgb='00CCCCFF') #good
    style06 = openpyxl.styles.colors.Color(rgb='00FF99CC') #inco

    style1 = PatternFill(patternType='solid', fgColor=style01)
    style2 = PatternFill(patternType='solid', fgColor=style02)
    style3 = PatternFill(patternType='solid', fgColor=style03)
    style4 = PatternFill(patternType='solid', fgColor=style04)
    style5 = PatternFill(patternType='solid', fgColor=style05)
    style6 = PatternFill(patternType='solid', fgColor=style06)
    
    imputation_method_id = int(imputation_method_id)
    if(imputation_method_id == 1):
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            if(var_type_final_2[j] == 'categorical'):
                v = np.around(my_data[:,j][:])
                if(np.max(v)>np.max(my_data[:,j][:])):
                    v = np.floor(my_data[:,j][:])
                my_data[:,j] = v
    elif(imputation_method_id == 2):
        T = np.isnan(data_org)
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            f = np.where(T[:,j] == False)
            f_nan = np.where(T[:,j] == True)
            X_f = copy.deepcopy(data_org.X[:,j])
            b = X_f[f]
            if(np.size(np.where(T[:,j] == 1)) <= np.size(b)):
                f_ind = random.sample(range(0, np.size(b)), np.size(np.where(T[:,j] == 1)))
                X_f[f_nan] = b[f_ind]
                my_data[:,j] = X_f.reshape(-1,1)
            else:
                my_data[:,j] = X_f.reshape(-1,1)
    else:
        my_data = data_org
                        
    c1 = 0
    c2 = 0
    
    sheet_names2 = wb2.sheetnames
    sheet = wb2[sheet_names2[0]]
    
    for j in range(c+len(pos_metas)):
        if(j not in pos_metas):
            if(features_state[c1] == 'bad'):
                sheet1.cell(1,j+1,features_total[j])
                sheet1.cell(1,j+1).fill = style3
            elif(features_state[c1] == 'fair'):
                sheet1.cell(1,j+1,features_total[j])
                sheet1.cell(1,j+1).fill = style4
            elif(features_state[c1] == 'good'):
                sheet1.cell(1,j+1,features_total[j])
                sheet1.cell(1,j+1).fill = style5
            # print("Feature ", str(j))
            for i in range(r):
                # print("Feature ", str(j), ", row ", str(i), ", value", str(data_org[:,c1][i]))
                if(outliers_pos[c1] != '-'):
                    if((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(my_data[:,c1][i]).strip()[1:-1]))
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')&(var_type_final[c1]!='string')&(var_type_final[c1]!='unknown')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                        sheet1.cell(i+2,j+1).fill = style2
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(my_data[:,c1][i]).strip()[1:-1]))
                        sheet1.cell(i+2,j+1).fill = style2
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                else:
                    # print(str(c1)+","+str(i))
                    try:
                        if((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')&(var_type_final[c1]!='string')&(var_type_final[c1]!='unknown')):
                            sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                            sheet1.cell(i+2,j+1).fill = style2
                        elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')):
                            sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                            sheet1.cell(i+2,j+1).fill = style2
                        elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                            if(incomp_pos[c1] == '-'):
                                sheet1.cell(i+2,j+1,formatNumber_v3(str(my_data[:,c1][i]).strip()[1:-1]))
                                sheet1.cell(i+2,j+1).fill = style2
                            else:
                                sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                                sheet1.cell(i+2,j+1).fill = style2
                        elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]=='bad')):
                            if(i in list(incomp_pos[c1])):
                                sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                                sheet1.cell(i+2,j+1).fill = style6
                            else:
                                sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                        elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                            if(i in list(incomp_pos[c1])):
                                sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                                sheet1.cell(i+2,j+1).fill = style6
                            else:
                                sheet1.cell(i+2,j+1,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                    except:
                        sheet1.cell(i+2,j+1,'?')
                        sheet1.cell(i+2,j+1).fill = style6
            c1 = c1+1
        elif(j in pos_metas):
            if(features_state_metas[c2] == 'bad'):
                sheet1.cell(1,j+1,features_total[j])
                sheet1.cell(1,j+1).fill = style3
            elif(features_state_metas[c2] == 'fair'):
                sheet1.cell(1,j+1,features_total[j])
                sheet1.cell(1,j+1).fill = style4
            elif(features_state_metas[c2] == 'good'):
                sheet1.cell(1,j+1,features_total[j])
                sheet1.cell(1,j+1).fill = style5
            for i in range(r):
                if(outliers_pos_metas[c2] != '-'):
                    if((i in outliers_pos_metas[c2])&(str(sheet.cell(i+2,j+1).value)=='')&(features_state_metas[c2]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str('?').strip()))    
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((i in outliers_pos_metas[c2])&(str(sheet.cell(i+2,j+1).value)=='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((i in outliers_pos_metas[c2])&(str(sheet.cell(i+2,j+1).value)!='')&(features_state_metas[c2]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((i in outliers_pos_metas[c2])&(str(sheet.cell(i+2,j+1).value)!='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))    
                        sheet1.cell(i+2,j+1).fill = style1
                    elif((str(sheet.cell(i+1,j+1).value).strip()=='')&(features_state_metas[c2]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,j+1).fill = style2
                    elif((str(sheet.cell(i+1,j+1).value).strip()=='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,j+1).fill = style2
                    elif((str(sheet.cell(i+1,j+1).value).strip()!='')&(features_state_metas[c2]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))
                    elif((str(sheet.cell(i+1,j+1).value).strip()!='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))
                else:
                    if((str(sheet.cell(i+1,j+1).value).strip()=='')&(features_state_metas[c2]=='bad')&(var_type_metas[c2]!='string')&(var_type_metas[c2]!='unknown')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,j+1).fill = style2
                    elif((str(sheet.cell(i+1,j+1).value).strip()=='')&(features_state_metas[c2]=='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str('?').strip())) 
                        sheet1.cell(i+2,j+1).fill = style2
                    elif((str(sheet.cell(i+1,j+1).value).strip()=='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,j+1,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,j+1).fill = style2
                    elif((str(sheet.cell(i+1,j+1).value).strip()!='')&(features_state_metas[c2]=='bad')):
                        if(i in list(incomp_pos_metas[c2])):
                            sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))
                            sheet1.cell(i+2,j+1).fill = style6
                        else:
                            sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))
                    elif((str(sheet.cell(i+1,j+1).value).strip()!='')&(features_state_metas[c2]!='bad')):
                        if(i in list(incomp_pos_metas[c2])):
                            sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))
                            sheet1.cell(i+2,j+1).fill = style6
                        else:
                            sheet1.cell(i+2,j+1,formatNumber_v3(str(sheet.cell(i+2,j+1).value).strip()))
            c2 = c2+1
            
    if not os.path.exists('results'):
        os.makedirs('results')
    
    wb.save(path_f)


def write_curated_dataset_dict(data_org, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                            var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                            var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, var_type_final_2, T_data_org, ncols, nrows):
    
    imputation_method_id = int(imputation_method_id)
    if(imputation_method_id == 1):
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            if(var_type_final_2[j] == 'categorical'):
                v = np.around(my_data[:,j][:])
                if(np.max(v)>np.max(my_data[:,j][:])):
                    v = np.floor(my_data[:,j][:])
                my_data[:,j] = v
    elif(imputation_method_id == 2):
        T = np.isnan(data_org)
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            f = np.where(T[:,j] == False)
            f_nan = np.where(T[:,j] == True)
            X_f = copy.deepcopy(data_org.X[:,j])
            b = X_f[f]
            if(np.size(np.where(T[:,j] == 1)) <= np.size(b)):
                f_ind = random.sample(range(0, np.size(b)), np.size(np.where(T[:,j] == 1)))
                X_f[f_nan] = b[f_ind]
                my_data[:,j] = X_f.reshape(-1,1)
            else:
                my_data[:,j] = X_f.reshape(-1,1)
    else:
        my_data = data_org
        
    c1 = 0
    c2 = 0
    
    features_states = []
    write_features = []
    styless = []
    styless_f = []
    
    for j in range(c+len(pos_metas)):
        if(j not in pos_metas):
            if(features_state[c1] == 'bad'):
                features_states.append(features_total[j]); styless_f.append('red')
            elif(features_state[c1] == 'fair'):
                features_states.append(features_total[j]); styless_f.append('blue')
            elif(features_state[c1] == 'good'):
                features_states.append(features_total[j]); styless_f.append('green')
            for i in range(r):
                if(outliers_pos[c1] != '-'):
                    if((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('orange')
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        write_features.append(formatNumber_v3(str(my_data[:,c1][i]).strip()[1:-1])); styless.append('orange')
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]=='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('orange')
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('orange')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')&(var_type_final[c1]!='string')&(var_type_final[c1]!='unknown')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('gray')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        write_features.append(formatNumber_v3(str(my_data[:,c1][i]).strip()[1:-1])); styless.append('gray')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]=='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('white')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('white')
                else:
                    if((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')&(var_type_final[c1]!='string')&(var_type_final[c1]!='unknown')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('gray')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]=='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('gray')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        if(incomp_pos[c1] == '-'):
                            write_features.append(formatNumber_v3(str(my_data[:,c1][i]).strip()[1:-1])); styless.append('gray')
                        else:
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('gray')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]=='bad')):
                        if(i in list(incomp_pos[c1])):
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('red')
                        else:
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('white')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        if(i in list(incomp_pos[c1])):
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('red')
                        else:
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('white')           
            c1 = c1+1
        elif(j in pos_metas):
            if(features_state_metas[c2] == 'bad'):
                features_states.append(features_total[j]); styless_f.append('red')
            elif(features_state_metas[c2] == 'fair'):
                features_states.append(features_total[j]); styless_f.append('blue')
            elif(features_state_metas[c2] == 'good'):
                features_states.append(features_total[j]); styless_f.append('green')
            for i in range(r):
                if(outliers_pos_metas[c2] != '-'):
                    if((i in outliers_pos_metas[c2])&(str(T_data_org.iloc[i,j])=='')&(features_state_metas[c2]=='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('orange')
                    elif((i in outliers_pos_metas[c2])&(str(T_data_org.iloc[i,j])=='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('orange')                      
                    elif((i in outliers_pos_metas[c2])&(str(T_data_org.iloc[i,j])!='')&(features_state_metas[c2]=='bad')):
                        write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j]).strip())); styless.append('orange')
                    elif((i in outliers_pos_metas[c2])&(str(T_data_org.iloc[i,j])!='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j]).strip())); styless.append('orange')
                    elif((str(T_data_org.iloc[i,j]).strip()=='')&(features_state_metas[c2]=='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('gray')           
                    elif((str(T_data_org.iloc[i,j]).strip()=='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('gray')             
                    elif((str(T_data_org.iloc[i,j]).strip()!='')&(features_state_metas[c2]=='bad')):
                        write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j].strip()))); styless.append('white')
                    elif((str(T_data_org.iloc[i,j]).strip()!='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j]).strip())); styless.append('white')
                else:
                    if((str(T_data_org.iloc[i,j]).strip()=='')&(features_state_metas[c2]=='bad')&(var_type_metas[c2]!='string')&(var_type_metas[c2]!='unknown')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('gray')
                    elif((str(T_data_org.iloc[i,j]).strip()=='')&(features_state_metas[c2]=='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('gray')
                    elif((str(T_data_org.iloc[i,j]).strip()=='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('gray')
                    elif((str(T_data_org.iloc[i,j]).strip()!='')&(features_state_metas[c2]=='bad')):
                        if(i in list(incomp_pos_metas[c2])):
                            write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j]).strip())); styless.append('red')
                        else:
                            write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j]).strip())); styless.append('white')
                    elif((str(T_data_org.iloc[i,j]).strip()!='')&(features_state_metas[c2]!='bad')):
                        if(i in list(incomp_pos_metas[c2])):
                            write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j]).strip())); styless.append('red')
                        else:
                            write_features.append(formatNumber_v3(str(T_data_org.iloc[i,j]).strip())); styless.append('white')
            c2 = c2+1
            
    write_features_v2 = np.array(write_features).reshape(ncols, nrows).tolist()
    styless_f_v2 = np.array(styless).reshape(ncols, nrows).tolist()

    python_dict = [{'Values from the curated dataset':write_features_v2,
                    'Style for feature names in the curated dataset': styless_f,
                    'Style for feature values in the curated dataset': styless_f_v2}]
    
    return python_dict


def write_curated_dataset_v2(data_org, wb2, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                             var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                             var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, path_f, var_type_final_2):
    
    wb = Workbook()
    sheet_names = wb.sheetnames
    sheet1 = wb[sheet_names[0]]
    
    # sheet1.set_column(1, np.size(data_org,1), 18)
    
    style01 = openpyxl.styles.colors.Color(rgb='00ECE338') #outlier
    style02 = openpyxl.styles.colors.Color(rgb='00D9D9D9') #imputed
    style04 = openpyxl.styles.colors.Color(rgb='00CCFFCC') #fair
    style05 = openpyxl.styles.colors.Color(rgb='00CCCCFF') #good
    style06 = openpyxl.styles.colors.Color(rgb='00FF99CC') #inco

    style1 = PatternFill(patternType='solid', fgColor=style01)
    style2 = PatternFill(patternType='solid', fgColor=style02)
    style4 = PatternFill(patternType='solid', fgColor=style04)
    style5 = PatternFill(patternType='solid', fgColor=style05)
    style6 = PatternFill(patternType='solid', fgColor=style06)
    
    imputation_method_id = int(imputation_method_id)
    if(imputation_method_id == 1):
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            if(var_type_final_2[j] == 'categorical'):
                v = np.around(my_data[:,j][:])
                if(np.max(v)>np.max(my_data[:,j][:])):
                    v = np.floor(my_data[:,j][:])
                my_data[:,j] = v
    elif(imputation_method_id == 2):
        T = np.isnan(data_org)
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            f = np.where(T[:,j] == False)
            f_nan = np.where(T[:,j] == True)
            X_f = copy.deepcopy(data_org.X[:,j])
            b = X_f[f]
            if(np.size(np.where(T[:,j] == 1)) <= np.size(b)):
                f_ind = random.sample(range(0, np.size(b)), np.size(np.where(T[:,j] == 1)))
                X_f[f_nan] = b[f_ind]
                my_data[:,j] = X_f.reshape(-1,1)
            else:
                my_data[:,j] = X_f.reshape(-1,1)
    else:
        my_data = data_org


    c1 = 0
    c2 = 0
    c3 = 0
    
    sheet_names2 = wb2.sheetnames
    sheet = wb2[sheet_names2[0]]
    
    for j in range(c+len(pos_metas)):
        if(j not in pos_metas):
            if(features_state[c1] == 'fair'):
                c3 = c3+1
                sheet1.cell(1,c3,features_total[j])
                sheet1.cell(1,c3).fill = style4
            elif(features_state[c1] == 'good'):
                c3 = c3+1
                sheet1.cell(1,c3,features_total[j])
                sheet1.cell(1,c3).fill = style5
            for i in range(r):
                if(outliers_pos[c1] != '-'):
                    if((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,c3,'?')
                        sheet1.cell(i+2,c3).fill = style1
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,c3,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                        sheet1.cell(i+2,c3).fill = style1
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,c3,'?')
                        sheet1.cell(i+2,c3).fill = style2
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        sheet1.cell(i+2,c3,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                else:
                    try:
                        if((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                            if(incomp_pos[c1] == '-'):
                                sheet1.cell(i+2,c3,'?')
                                sheet1.cell(i+2,c3).fill = style2
                            else:
                                sheet1.cell(i+2,c3,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                                sheet1.cell(i+2,c3).fill = style2
                        elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                            if(i in list(incomp_pos[c1])):
                                sheet1.cell(i+2,c3,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                                sheet1.cell(i+2,c3).fill = style6
                            else:
                                sheet1.cell(i+2,c3,formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1]))
                    except:
                        sheet1.cell(i+2,c3,'?')
                        sheet1.cell(i+2,c3).fill = style6
            c1 = c1+1
        elif(j in pos_metas):
            if(features_state_metas[c2] == 'fair'):
                c3 = c3+1
                sheet1.cell(1,c3,features_total[j])
                sheet1.cell(1,c3).fill = style4
            elif(features_state_metas[c2] == 'good'):
                c3 = c3+1
                sheet1.cell(1,c3,features_total[j])
                sheet1.cell(1,c3).fill = style5
            for i in range(r):
                if(outliers_pos_metas[c2] != '-'):                           
                    if((i in outliers_pos_metas[c2])&(str(sheet.cell(i+2,c3).value)=='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,c3,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,c3).fill = style1
                    elif((i in outliers_pos_metas[c2])&(str(sheet.cell(i+2,c3).value)!='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,c3,formatNumber_v3(str(sheet.cell(i+2,c3).value).strip())) 
                        sheet1.cell(i+2,c3).fill = style1
                    elif((str(sheet.cell(i+2,c3).value).strip()=='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,c3,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,c3).fill = style2
                    elif((str(sheet.cell(i+2,c3).value).strip()!='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,c3,formatNumber_v3(str(sheet.cell(i+2,c3).value).strip()))
                else:
                    if((str(sheet.cell(i+2,c3).value).strip()=='')&(features_state_metas[c2]!='bad')):
                        sheet1.cell(i+2,c3,formatNumber_v3(str('?').strip()))
                        sheet1.cell(i+2,c3).fill = style2
                    elif((str(sheet.cell(i+2,c3).value).strip()!='')&(features_state_metas[c2]!='bad')):
                        if(i in list(incomp_pos_metas[c2])):
                            sheet1.cell(i+2,c3,formatNumber_v3(str(sheet.cell(i+2,c3).value).strip()))
                            sheet1.cell(i+2,c3).fill = style6
                        else:
                            sheet1.cell(i+2,c3,formatNumber_v3(str(sheet.cell(i+2,c3).value).strip()))
            c2 = c2+1
    
    if not os.path.exists('results'):
        os.makedirs('results')
    
    wb.save(path_f)


def write_curated_dataset_v2_dict(data_org, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                                var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                                var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, var_type_final_2, T_data_org, ncols, nrows):
    
    imputation_method_id = int(imputation_method_id)
    if(imputation_method_id == 1):
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            if(var_type_final_2[j] == 'categorical'):
                v = np.around(my_data[:,j][:])
                if(np.max(v)>np.max(my_data[:,j][:])):
                    v = np.floor(my_data[:,j][:])
                my_data[:,j] = v
    elif(imputation_method_id == 2):
        T = np.isnan(data_org)
        imputer = Impute(method=Average())
        my_data = imputer(data_org)
        for j in range(c):
            f = np.where(T[:,j] == False)
            f_nan = np.where(T[:,j] == True)
            X_f = copy.deepcopy(data_org.X[:,j])
            b = X_f[f]
            if(np.size(np.where(T[:,j] == 1)) <= np.size(b)):
                f_ind = random.sample(range(0, np.size(b)), np.size(np.where(T[:,j] == 1)))
                X_f[f_nan] = b[f_ind]
                my_data[:,j] = X_f.reshape(-1,1)
            else:
                my_data[:,j] = X_f.reshape(-1,1)
    else:
        my_data = data_org
                        
    c1 = 0
    c2 = 0
    c3 = -1
    
    features_states = []
    write_features = []
    styless = []
    styless_f = []
    
    for j in range(c+len(pos_metas)):
        if(j not in pos_metas):
            if(features_state[c1] == 'fair'):
                c3 = c3+1
                features_states.append(features_total[j]); styless_f.append('fair')
            elif(features_state[c1] == 'good'):
                c3 = c3+1
                features_states.append(features_total[j]); styless_f.append('blue')
            for i in range(r):
                if(outliers_pos[c1] != '-'):
                    if((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        write_features.append('NaN'); styless.append('orange')
                    elif((i in outliers_pos[c1])&(str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('orange')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        write_features.append('NaN'); styless.append('gray')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('white')
                else:
                    if((str(data_org[:,c1][i]).strip()[1:-1]=='?')&(features_state[c1]!='bad')):
                        if(incomp_pos[c1] == '-'):
                            write_features.append('NaN'); styless.append('gray')
                        else:
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('gray')
                    elif((str(data_org[:,c1][i]).strip()[1:-1]!='?')&(features_state[c1]!='bad')):
                        if(i in list(incomp_pos[c1])):
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('red')
                        else:
                            write_features.append(formatNumber_v3(str(data_org[:,c1][i]).strip()[1:-1])); styless.append('white')       
            c1 = c1+1
        elif(j in pos_metas):
            if(features_state_metas[c2] == 'fair'):
                c3 = c3+1
                features_states.append(features_total[j]); styless_f.append('fair')
            elif(features_state_metas[c2] == 'good'):
                c3 = c3+1
                features_states.append(features_total[j]); styless_f.append('blue')
            for i in range(r):
                if(outliers_pos_metas[c2] != '-'):                    
                    if((i in outliers_pos_metas[c2])&(str(T_data_org.iloc[i,c3])=='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('orange')
                    elif((i in outliers_pos_metas[c2])&(str(T_data_org.iloc[i,c3])!='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str(T_data_org.iloc[i,c3]).strip())); styless.append('orange');                         
                    elif((str(T_data_org.iloc[i,c3]).strip()=='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('gray')
                    elif((str(T_data_org.iloc[i,c3]).strip()!='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str(T_data_org.iloc[i,c3]).strip())); styless.append('white')
                else:
                    if((str(T_data_org.iloc[i,c3]).strip()=='')&(features_state_metas[c2]!='bad')):
                        write_features.append(formatNumber_v3(str('NaN').strip())); styless.append('gray')
                    elif((str(T_data_org.iloc[i,c3]).strip()!='')&(features_state_metas[c2]!='bad')):
                        if(i in list(incomp_pos_metas[c2])):
                            write_features.append(formatNumber_v3(str(T_data_org.iloc[i,c3]).strip())); styless.append('red')
                        else:
                            write_features.append(formatNumber_v3(str(T_data_org.iloc[i,c3]).strip())); styless.append('white')
            c2 = c2+1
    

    write_features_v2 = np.array(write_features).reshape(ncols, nrows).tolist()
    styless_f_v2 = np.array(styless).reshape(ncols, nrows).tolist()
    python_dict = [{'Values from the clean curated dataset': write_features_v2,
                    'Style for feature names in the clean curated dataset': styless_f,
                    'Style for feature values in the clean curated dataset': styless_f_v2}]
    
    return python_dict


def QualityAssessment_S1(data_org, c, th):
    features_missing_values = []
    bad_features = []
    bad_features_ind = []
    fair_features = []
    fair_features_ind = []
    good_features = []
    good_features_ind = []
    features_state = []
    a_total = 0
    a = np.zeros(c)
        
    for i in range(c):
        a[i] = sum(1 for d in data_org if np.isnan(d[i]))
        features_missing_values.append(a[i])
        if(a[i]>=th):
            bad_features.append(data_org.domain.attributes[i].name.replace('\n',' ').replace('  ', ' '))
            features_state.append('bad') 
            bad_features_ind.append(i)
        elif((a[i]>0) & (a[i]<th)):
            fair_features.append(data_org.domain.attributes[i].name.replace('\n',' ').replace('  ', ' '))
            features_state.append('fair') 
            fair_features_ind.append(i)
        elif(a[i]==0):
            good_features.append(data_org.domain.attributes[i].name.replace('\n',' ').replace('  ', ' '))
            features_state.append('good') 
            good_features_ind.append(i)
        a_total = a_total+a[i]
    
    return [features_missing_values, bad_features, bad_features_ind, fair_features, fair_features_ind, 
            good_features, good_features_ind, features_state, a, a_total]


def outliers_detection(data_org, c, var_type_final, outlier_detection_method_id, bad_features_ind, fair_features_ind, good_features_ind, features_missing_values):
    T = np.isnan(data_org)
    outliers_ind = []
    y_score = []
    outliers_pos = []
    
    if(outlier_detection_method_id == 1):
        for j in range(c):
            if((var_type_final[j] == 'string')|(var_type_final[j] == 'unknown')):
                outliers_ind.append('not-applicable')
                y_score.append('-') 
                outliers_pos.append('-')
            elif(((var_type_final[j] == 'float')|(var_type_final[j] == 'date')|(var_type_final[j] == 'int'))&(features_missing_values[j] != len(data_org))):
                    f = np.where(T[:,j]==False)
                    X_f = data_org.X[:,j]
                    b = X_f[f]
                    [z_scores, outliers_ind_z_scores] = outliers_z_score(b)
                    y_score.append(np.mean(z_scores))
                    if(len(outliers_ind_z_scores[0]) != 0):
                        if(var_type_final[j] == 'int')&(np.min(b) == 0)&(np.max(b) == 1):
                            outliers_ind.append('no')
                            outliers_pos.append('-')
                        else:
                            t = list(outliers_ind_z_scores[0])
                            outliers_pos.append(list(f[0][t]))
                            outliers_ind.append('yes')
                    else:
                        outliers_ind.append('no')
                        outliers_pos.append('-')
            elif(((var_type_final[j] == 'float')|(var_type_final[j] == 'date')|(var_type_final[j] == 'int'))&(features_missing_values[j] == len(data_org))):
                outliers_ind.append('not-applicable')
                outliers_pos.append('-')
    elif(outlier_detection_method_id == 2):
        for j in range(c):
            if((var_type_final[j] == 'string')|(var_type_final[j] == 'unknown')):
                outliers_ind.append('not-applicable')
                y_score.append('-') 
                outliers_pos.append('-')
            elif(((var_type_final[j] == 'float')|(var_type_final[j] == 'date')|(var_type_final[j] == 'int'))&(features_missing_values[j] != len(data_org))):
                    f = np.where(T[:,j]==False)
                    X_f = data_org.X[:,j]
                    b = X_f[f]
                    [z_scores, outliers_ind_z_scores] = outliers_z_score_mod(b)
                    y_score.append(np.mean(z_scores))
                    if(len(outliers_ind_z_scores[0]) != 0):
                        if(var_type_final[j] == 'int')&(np.min(b) == 0)&(np.max(b) == 1):
                            outliers_ind.append('no')
                            outliers_pos.append('-')
                        else:
                            t = list(outliers_ind_z_scores[0])
                            outliers_pos.append(list(f[0][t]))
                            outliers_ind.append('yes')
                    else:
                        outliers_ind.append('no')
                        outliers_pos.append('-')
            elif(((var_type_final[j] == 'float')|(var_type_final[j] == 'date')|(var_type_final[j] == 'int'))&(features_missing_values[j] == len(data_org))):
                outliers_ind.append('not-applicable')
                outliers_pos.append('-')
    elif(outlier_detection_method_id == 3):
        for j in range(c):
            if((var_type_final[j] == 'string')|(var_type_final[j] == 'unknown')):
                outliers_ind.append('not-applicable')
                y_score.append('-') 
                outliers_pos.append('-')
            elif(((var_type_final[j] == 'float')|(var_type_final[j] == 'date')|(var_type_final[j] == 'int'))&(features_missing_values[j] != len(data_org))):
                    f = np.where(T[:,j]==False)
                    X_f = data_org.X[:,j]
                    b = X_f[f]
                    [iqr, outliers_ind_iqr] = outliers_iqr(b)
                    y_score.append(np.mean(iqr))
                    if(len(outliers_ind_iqr[0]) != 0):
                        if(var_type_final[j] == 'int')&(np.min(b) == 0)&(np.max(b) == 1):
                            outliers_ind.append('no')
                            outliers_pos.append('-')
                        else:
                            t = list(outliers_ind_iqr[0])
                            outliers_pos.append(list(f[0][t]))
                            outliers_ind.append('yes')
                    else:
                        outliers_ind.append('no')
                        outliers_pos.append('-')
            elif(((var_type_final[j] == 'float')|(var_type_final[j] == 'date')|(var_type_final[j] == 'int'))&(features_missing_values[j] == len(data_org))):
                outliers_ind.append('not-applicable')
                outliers_pos.append('-')
    elif(outlier_detection_method_id == 4):
        for j in range(c):
            if((var_type_final[j] == 'string')|(var_type_final[j] == 'unknown')):
                outliers_ind.append('not-applicable')
                y_score.append('-') 
                outliers_pos.append('-')
            elif(((var_type_final[j] == 'float')|(var_type_final[j] == 'date')|(var_type_final[j] == 'int'))&(features_missing_values[j] != len(data_org))):
                    f = np.where(T[:,j]==False)
                    X_f = data_org.X[:,j]
                    b = X_f[f]
                    h_min = grubbs.min_test_indices(b)
                    h_max = grubbs.max_test_indices(b)
                    h_tol = np.unique(np.union1d(h_min, h_max))
                    y_score.append(np.mean(grubbs.test(b, alpha=0.05)))
                    if(len(h_tol) != 0):
                        if(var_type_final[j] == 'int')&(np.min(b) == 0)&(np.max(b) == 1):
                            outliers_ind.append('no')
                            outliers_pos.append('-')
                        else:
                            t = list(h_tol)
                            outliers_pos.append(list(np.take(f,t)))
                            outliers_ind.append('yes')
                    else:
                        outliers_ind.append('no')
                        outliers_pos.append('-')
    # elif(outlier_detection_method_id == 5):
    #     X_c = data_org.X[:,good_features_ind] #IMPORTANT
    #     # clusterer = hdbscan.HDBSCAN(min_cluster_size=2).fit(X_c)
    #     y_pred_lof = clusterer.outlier_scores_ #lof_ind = np.where(y_score > 0.85)[0]
    #     y_score.append(np.mean(y_pred_lof))
        
    #     c1 = 0
    #     for j in range(c):
    #         if(j in good_features_ind):            
    #             if(y_pred_lof[c1] >= 0.8):
    #                 outliers_ind.append('yes')
    #                 outliers_pos.append('-')
    #             else:
    #                 outliers_ind.append('no')
    #                 outliers_pos.append('-')
    #             c1 = c1+1
    #         else:
    #             outliers_ind.append('not-applicable')
    #             outliers_pos.append('-')
    elif(outlier_detection_method_id == 6):
        model = IsolationForest(n_estimators=100, contamination=0.02)
        X_c = data_org.X[:,good_features_ind] #IMPORTANT
        
        #idea 1
        model.fit(X_c)
        anomaly = model.predict(X_c)
        out_ind = np.where(anomaly == -1)[0]
        
        for j in range(c):
            if(j in good_features_ind):
                if(len(out_ind) != 0):
                    outliers_ind.append('yes')
                    outliers_pos.append(out_ind)
                else:
                    outliers_ind.append('no')
                    outliers_pos.append('-')                   
            else:
                outliers_ind.append('not-applicable')
                outliers_pos.append('-')
    elif(outlier_detection_method_id == 7):
        model = IsolationForest(n_estimators=100, contamination=0.01)
        X_c = data_org.X[:,good_features_ind] #IMPORTANT
                
        #idea 2        
        c1 = 0
        for j in range(c):
            if(j in good_features_ind):
                model.fit(X_c[:,c1].reshape(-1,1))
                anomaly = model.predict(X_c[:,c1].reshape(-1,1))
                if(-1 in anomaly):
                    outliers_ind.append('yes')
                    outliers_pos.append(list(np.where(anomaly == -1)[0]))
                else:
                    outliers_ind.append('no')
                    outliers_pos.append('-')
                c1 = c1+1
            else:
                outliers_ind.append('not-applicable')
                outliers_pos.append('-')                 
    else:
        for j in range(c):
            outliers_pos.append('-')
            outliers_ind.append('no')
            y_score = []         
            
    return [outliers_ind, y_score, outliers_pos]


def metas_handler(data_org, features, outlier_detection_method_id, T_data_org):
    metas = data_org.domain.metas
    metas_features = []
    for i in range(len(metas)):
        metas_features.append(metas[i].name)

    features_total = []
    number_of_columns = np.size(data_org.X, 1)
    features_total = features

    pos_metas = [];
    for i in range(len(features_total)):
        for j in range(len(metas_features)):
            if(jaro(metas_features[j], features_total[i]) == 1):
                pos_metas.append(i)
            
    number_of_rows = np.size(data_org.X, 0)
    y_total = []
    var_type_metas = []
    var_type_metas_2 = []
    incompatibilities_metas = []
    features_state_metas = []
    ranges_metas = []
    means_metas = []
    incomp_pos_metas = []
    
    #data annotation
    for j in range(len(pos_metas)):
        y_r = []
        for i in range(1,number_of_rows):
            y = T_data_org.iloc[i,pos_metas[j]]; #I THINK IT NEEDS DATAFRAME HERE <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
            try:
                y_mod = float(y)
                y_mod = formatNumber(y_mod)
                y_r.append(y_mod)
            except:
                y_r.append(y)
        y_total.append(y_r)
    
    for k in range(np.size(y_total,0)):
        types = []
        for l in range(np.size(y_total,1)):
            types.append((type(y_total[k][l])))
        matches_str = [1 for x in types if x==str]
        matches_int = [1 for x in types if x==int]
        matches_float = [1 for x in types if x==float]
        if(len(matches_str) == len(types)):
            var_type_metas.append('string')
            var_type_metas_2.append('categorical')
            incompatibilities_metas.append('no'); means_metas.append('-')
            ranges_metas.append(list(set(y_total[k]))); incomp_pos_metas.append('-')
        elif(len(matches_int) == len(types)):
            if('year' in metas_features[j])|('Date' in metas_features[j])|('date' in metas_features[j])|('year' in metas_features[j])|('yr' in metas_features[j])|('Dates' in metas_features[j])|('Year' in metas_features[j])|('YEAR' in metas_features[j]):
                var_type_metas.append('date'); means_metas.append(np.median(y_total[k]))
            else:
                var_type_metas.append('int'); means_metas.append(np.mean(y_total[k]))
            incompatibilities_metas.append('no')
            a = np.str(formatNumber(np.nanmin(y_total[k])))+','+np.str(formatNumber(np.nanmax(y_total[k])))
            ranges_metas.append(a.split(',')); incomp_pos_metas.append('-')
            
            #NEW condition
            if((len(str(np.nanmax(y_total[k])))) == 1)&((len(str(np.nanmin(y_total[k])))) == 1):
                var_type_metas_2.append('categorical')
            else:
                var_type_metas_2.append('numeric')
            
        elif(len(matches_float) == len(types)):
            if('year' in metas_features[j])|('Date' in metas_features[j])|('date' in metas_features[j])|('year' in metas_features[j])|('yr' in metas_features[j])|('Dates' in metas_features[j])|('Year' in metas_features[j])|('YEAR' in metas_features[j]):
                var_type_metas.append('date'); means_metas.append(np.median(y_total[k]))
            else:
                var_type_metas.append('float'); means_metas.append(np.mean(y_total[k]))
            var_type_metas_2.append('numeric')
            incompatibilities_metas.append('no')
            a = np.str(formatNumber(np.nanmin(y_total[k])))+','+np.str(formatNumber(np.nanmax(y_total[k])))
            ranges_metas.append(a.split(',')); incomp_pos_metas.append('-')
        elif((len(matches_int)+len(matches_float) == len(types))&(len(matches_float)!=0)&(len(matches_int)!=0)):
            var_type_metas.append('float'); means_metas.append(np.mean(y_total[k]))
            var_type_metas_2.append('numeric')
            incompatibilities_metas.append('no')
            a = np.str(formatNumber(np.nanmin(y_total[k])))+','+np.str(formatNumber(np.nanmax(y_total[k])))
            ranges_metas.append(a.split(',')); incomp_pos_metas.append('-');          
        else:
            var_type_metas.append('unknown'); means_metas.append('-')
            var_type_metas_2.append('unknown')
            incompatibilities_metas.append("yes, unknown type of data")
            ranges_metas.append(list(set(y_total[k])))
            p = [i for i,x in enumerate(types) if ((x==str)&(np.str(y_total[k][i]).strip()!=''))]
            incomp_pos_metas.append(p)
        
    #missing values
    features_missing_values_metas = []
    bad_features_metas = []
    bad_features_ind_metas = []
    fair_features_metas = []
    fair_features_ind_metas = []
    good_features_metas = []
    good_features_ind_metas = []
    features_state_metas = []
    a_total_metas = 0
        
    for j in range(len(pos_metas)):
        c1 = 0
        for index, s in enumerate(y_total[j]):
            if(s == ''):
                c1 = c1+1;       
        features_missing_values_metas.append(c1)
        if(c1>=(number_of_rows*0.7)):
            bad_features_metas.append(metas_features[j])
            bad_features_ind_metas.append(j)
            features_state_metas.append('bad')
        elif((c1>0)&(c1<(number_of_rows*0.7))):
            fair_features_metas.append(metas_features[j])
            fair_features_ind_metas.append(j)
            features_state_metas.append('fair')   
        elif(c1==0):
            good_features_metas.append(metas_features[j])
            good_features_ind_metas.append(j)
            features_state_metas.append('good')
        a_total_metas = a_total_metas+c1
    
    #outlier detection
    outliers_ind_metas = []
    y_score_metas = []
    outliers_pos_metas = []
    if(outlier_detection_method_id == 1):
        for j in range(len(pos_metas)):
            if((var_type_metas[j] == 'string')|(var_type_metas[j] == 'unknown')):
                outliers_ind_metas.append('not-applicable')
                y_score_metas.append('-'); outliers_pos_metas.append('-')
            elif(((var_type_metas[j] == 'float')|(var_type_metas[j] == 'date')|(var_type_metas[j] == 'int'))&(features_missing_values_metas[j]!=number_of_rows)):
                y_f = []
                for index, s in enumerate(y_total[j]):
                    if((s!='')&(isinstance(s,str)==False)):
                        y_f.append(s)
                    b = np.asarray(y_f)
                    [z_scores, outliers_ind_z_scores] = outliers_z_score(b)
                    y_score_metas.append(np.mean(z_scores))
                    if(len(outliers_ind_z_scores[0]) != 0):
                        if(var_type_metas[j] == 'int')&(np.max(b)==1)&(np.min(b)==0):
                            outliers_ind_metas.append('no')
                            outliers_pos_metas.append('-')
                        else:
                            outliers_pos_metas.append(list(outliers_ind_z_scores[0]))
                            outliers_ind_metas.append('yes')
                    else:
                        outliers_ind_metas.append('no')
                        outliers_pos_metas.append('-')
            elif(((var_type_metas[j] == 'float')|(var_type_metas[j] == 'date')|(var_type_metas[j] == 'int'))&(features_missing_values_metas[j]==number_of_rows)):
                outliers_ind_metas.append('no')
                outliers_pos_metas.append('-')
    elif(outlier_detection_method_id == 2):
        for j in range(len(pos_metas)):
            if((var_type_metas[j] == 'string')|(var_type_metas[j] == 'unknown')):
                outliers_ind_metas.append('not-applicable')
                y_score_metas.append('-'); outliers_pos_metas.append('-')
            elif((var_type_metas[j] == 'float')|(var_type_metas[j] == 'date')|(var_type_metas[j] == 'int')):
                y_f = []
                for index, s in enumerate(y_total[j]):
                    if((s!='')&(isinstance(s,str)==False)):
                        y_f.append(s)
                    b = np.asarray(y_f)
                    [iqr, outliers_ind_iqr] = outliers_iqr(b)
                    y_score_metas.append(np.mean(iqr))
                    if(len(outliers_ind_iqr[0]) != 0):
                        if(var_type_metas[j] == 'int')&(np.max(b)==1)&(np.min(b)==0):
                            outliers_ind_metas.append('no')
                            outliers_pos_metas.append('-')
                        else:
                            outliers_pos_metas.append(list(outliers_ind_iqr[0]))
                            outliers_ind_metas.append('yes')
                    else:
                        outliers_ind_metas.append('no')
                        outliers_pos_metas.append('-')
            elif(((var_type_metas[j] == 'float')|(var_type_metas[j] == 'date')|(var_type_metas[j] == 'int'))&(features_missing_values_metas[j]==number_of_rows)):
                outliers_ind_metas.append('no')
                outliers_pos_metas.append('-')
    elif(outlier_detection_method_id == 3):
        for j in range(len(pos_metas)):
            if((var_type_metas[j] == 'string')|(var_type_metas[j] == 'unknown')):
                outliers_ind_metas.append('not-applicable')
                y_score_metas.append('-'); outliers_pos_metas.append('-')
            elif((var_type_metas[j] == 'float')|(var_type_metas[j] == 'date')|(var_type_metas[j] == 'int')):
                y_f = []
                for index, s in enumerate(y_total[5]):
                    if((s!='')&(isinstance(s,str)==False)):
                        y_f.append(s)
                    b = np.asarray(y_f)
                    h_min = grubbs.min_test_indices(b)
                    h_max = grubbs.max_test_indices(b)
                    h_tol = np.unique(np.union1d(h_min, h_max))
                    y_score_metas.append(np.mean(grubbs.test(b, alpha=0.05)))
                    if(len(h_tol) != 0):
                        if(var_type_metas[j] == 'int')&(np.max(b)==1)&(np.min(b)==0):
                            outliers_ind_metas.append('no')
                            outliers_pos_metas.append('-')
                        else:
                            outliers_pos_metas.append(list(h_tol))
                            outliers_ind_metas.append('yes')
                    else:
                        outliers_ind_metas.append('no')
                        outliers_pos_metas.append('-')
            elif(((var_type_metas[j] == 'float')|(var_type_metas[j] == 'date')|(var_type_metas[j] == 'int'))&(features_missing_values_metas[j]==number_of_rows)):
                outliers_ind_metas.append('no')
                outliers_pos_metas.append('-')
    elif(outlier_detection_method_id == 4):
        c1 = 0;
        for j in range(len(pos_metas)):
            outliers_ind_metas.append('not-applicable')
            outliers_pos_metas.append('-')
    else:
        for j in range(len(pos_metas)):  
            outliers_pos_metas.append('-')
            outliers_ind_metas.append('no')
                        
    return [features_total, metas_features, pos_metas, y_total, var_type_metas, var_type_metas_2, 
            features_state_metas, incompatibilities_metas, features_missing_values_metas, bad_features_metas, 
            bad_features_ind_metas,fair_features_metas, fair_features_ind_metas, good_features_metas,
            good_features_ind_metas,a_total_metas, outliers_ind_metas, y_score_metas, outliers_pos_metas, 
            ranges_metas, incomp_pos_metas, means_metas]


def similarity_detection(wb, method):
    sheet_names = wb.sheetnames
    xl = wb[sheet_names[0]]
    
    ncols = xl.max_column
    nrows = xl.max_row
    cmatrix = np.zeros((ncols,ncols))
    pmatrix = np.zeros((ncols,ncols))
    
    features_total = []
    for col in range(ncols):
        features_total.append((xl.cell(1,col+1).value))
    
    mylist2 = []
    for col in range(ncols):
        mylist = []
        for row in range(nrows):
            mylist.append((xl.cell(row+2,col+1).value))
        mylist2.append(mylist)
    
    for j in range(ncols):
        a = mylist2[j]
        for k in range(j+1,ncols):
            b = mylist2[k]
            try:
                a = pd.to_numeric(a, errors='coerce')
                b = pd.to_numeric(b, errors='coerce')
                if(np.sum(np.isnan(a)) < xl.max_row/2)&(np.sum(np.isnan(b)) < xl.max_row/2):
                    if(method == 1):
                        # print("Running Spearman")
                        [c,p] = spearmanr(a,b,nan_policy='omit')
                    elif(method == 2):
                        # print("Running Pearson")
                        [c,p] = pearsonr(a,b)
                    elif(method == 3):
                        # print("Running Kendall")
                        [c,p] = kendalltau(a,b,nan_policy='omit')
                    elif(method == 4):
                        # print("Running Covariance")
                        c = np.cov(a, b, bias=False)[0,1]
                        p = 1
                    elif(method == 5):
                        # print("None")
                        c = 0
                        p = 1
                else:
                    c = 0 
                    p = 1
            except:
                c = 0 
                p = 1
            cmatrix[j,k] = c
            pmatrix[j,k] = p
            
    cmatrix = cmatrix + cmatrix.T
    pmatrix = pmatrix + pmatrix.T
    np.fill_diagonal(cmatrix,1)
    np.fill_diagonal(pmatrix,0)
    fnans = np.isnan(cmatrix)
    pnans = np.isnan(pmatrix)
    cmatrix[fnans] = 0
    pmatrix[pnans] = 100
    
    if(method != 4):
        f_cmatrix = np.where((np.tril(cmatrix,-1) >= 0.9) & (np.tril(cmatrix,-1) <= 1))
    else:
        f_cmatrix = np.where((np.tril(cmatrix,-1) >= 0.9) | (np.tril(cmatrix,-1) <= -1))
    
    f_cmatrix_names = []
    r_values = []
    p_values = []
    for k in range(np.size(f_cmatrix,1)):
        f_cmatrix_names.append('('+features_total[f_cmatrix[0][k]]+','+features_total[f_cmatrix[1][k]]+')')
        r_values.append(cmatrix[f_cmatrix[0][k], f_cmatrix[1][k]])
        p_values.append(pmatrix[f_cmatrix[0][k], f_cmatrix[1][k]])
    
    total_features_clean = []
    regex = re.compile('\(.+?\)')
    for i in range(ncols):
        total_features_clean.append(regex.sub('', features_total[i]))
        s = total_features_clean[i]
        if(s[-1] == ' '):
            total_features_clean[i] = s[0:len(s)-1]
    
    jmatrix = np.zeros((ncols,ncols))
    for m in range(ncols):
        for n in range(m+1,ncols):
            jdist = jaro(total_features_clean[m], total_features_clean[n])
            jmatrix[m,n] = jdist
    
    jmatrix = jmatrix + jmatrix.T       
    f_jmatrix = np.where((np.tril(jmatrix,-1) >= 0.92))
    f_jmatrix_names = []
    j_values = []
    for k in range(np.size(f_jmatrix,1)):
        f_jmatrix_names.append('('+features_total[f_jmatrix[0][k]]+','+features_total[f_jmatrix[1][k]]+')')
        j_values.append(jmatrix[f_jmatrix[0][k], f_jmatrix[1][k]])
    
    f1c = []
    f2c = []
    for s in range(len(f_cmatrix_names)):
        f1c.append(f_cmatrix_names[s].split(',')[0].replace('(',''))
        f2c.append(f_cmatrix_names[s].split(',')[1].replace(')',''))

    D_c = pd.DataFrame(data={'f1':f1c,
                             'f2':f2c,
                             'value':r_values})

    f1j = []
    f2j = []
    for s in range(len(f_jmatrix_names)):
        f1j.append(f_jmatrix_names[s].split(',')[0].replace('(',''))
        f2j.append(f_jmatrix_names[s].split(',')[1].replace(')',''))
    D_j = pd.DataFrame(data={'f1':f1j,
                             'f2':f2j,
                             'value':j_values})
    
    return [cmatrix, f_cmatrix, pmatrix, f_cmatrix_names, r_values, p_values, jmatrix, 
            f_jmatrix, f_jmatrix_names, j_values, features_total, D_c, D_j]


@app.route("/features/", methods = ['GET', 'POST'])
def get_feature_values():
    path = 'data/UoA_small/demo-pSS-code-v3.xlsx'
    wb = open_workbook(path)
    sheet_names = wb.sheet_names()
    xl = wb.sheet_by_name(sheet_names[0])
    f = xl.row_values(0)
    f = [x.replace('\n','') for x in f]
    
    feature_id = request.args.get('feature_id', type=int)
    if(feature_id > np.size(f))|(feature_id <= 0):
        return json.dumps({'Error':'Selected feature is out of bounds!'})
    
    b0 = xl.col_values(feature_id-1, 1)
#    b = list(filter(None, b0))
    sel_f = f[feature_id-1]
    sel_f_values = str(b0).replace('\n','')
    
    b1 = xl.col_values(np.size(f)-1, 1)
    sel_t_values = str(b1).replace('\n','')
    dictionary = [{'Selected feature':sel_f,
                   'Selected feature values':sel_f_values,
                   'Target values':sel_t_values}]
    
    d = create_wr_io('results_feature.txt', dictionary)
    
    return jsonify(d)


@app.route("/features/names/", methods = ['GET', 'POST'])
def get_features_names():
    path = 'data/UoA_small/demo-pSS-code-v3.xlsx'
    wb = open_workbook(path)
    sheet_names = wb.sheet_names()
    xl = wb.sheet_by_name(sheet_names[0])
    f = xl.row_values(0)
    f = [x.replace('\n','') for x in f]

    dictionary = [{'Features': f}]
    
    return jsonify(dictionary)


def jsonify_data_curator(path, imputation_method_id, outlier_detection_method_id, descriptive_feature_id, wb2, bad_features_ind, bad_features_ind_metas, metas_features, cmatrix, f_cmatrix):
    outlier_detection_methods = ['z-score', 'IQR', 'Grubbs', 'None']
    imputation_methods = ['mean', 'random', 'None']
    
    if(imputation_method_id == None):
        imputation_method_id = 3
    
    if(outlier_detection_method_id == None):
        outlier_detection_method_id = 4
        
    wb = open_workbook(path)
    sheet_names = wb.sheet_names()
    xl = wb.sheet_by_name(sheet_names[0])
    
    nof = xl.cell(1,4).value
    noi = xl.cell(2,4).value
    df = xl.cell(3,4).value
    cf = xl.cell(4,4).value
    un = xl.cell(5,4).value
    mv = xl.cell(6,4).value
    f = xl.col_values(0, 10)
    f = [f[i].replace('\n','')  for i,x in enumerate(f)]
    
    ranges = xl.col_values(5, 10)
    t = xl.col_values(9, 10)
    t2 = xl.col_values(11, 10)
    mvf = xl.col_values(13, 10)
    s = xl.col_values(16, 10)
    o = xl.col_values(18, 10)
    inco = xl.col_values(20, 10)
    
#    wb2 = open_workbook(path)
    sheet_names2 = wb2.sheet_names()
    xl2 = wb2.sheet_by_name(sheet_names2[0])
    
    if(descriptive_feature_id != 0):
        m = []
        med = []
        std = []
        sk = []
        kurt = []
        
        b0 = xl2.col_values(descriptive_feature_id-1, 1)
        b = list(filter(None, b0))
        
        try:
            m.append(np.around(np.mean(b), 2))
        except:
            m.append('None')

        try:
            med.append(np.around(np.median(b), 2))
        except:
            med.append('None')

        try:
            std.append(np.around(np.std(b), 2))
        except:
            std.append('None')

        try:
            sk.append(np.around(scipy.stats.skew(b), 2))
        except:
            sk.append('None')

        try:
            kurt.append(np.around(scipy.stats.kurtosis(b), 2))
        except:
            kurt.append('None')
    
    bf_ind = [i for i,x in enumerate(s) if x=='bad']
    bf_names = [f[i] for i in np.asarray(bf_ind,int)]
    
#    with Capturing() as output:
#        for i in range(np.size(f_cmatrix,1)):
#            print("(", repr(f[f_cmatrix[0][i]].replace('\n',' ')), ",", 
#                    repr(f[f_cmatrix[1][i]].replace('\n',' ')), ",", 
#                    repr(str(np.around(cmatrix[f_cmatrix[0][i], f_cmatrix[1][i]], 2))), ")")
    
#    with Capturing() as output:
    output = []
    for i in range(np.size(f_cmatrix,1)):
        output.append(["(" + repr(f[f_cmatrix[0][i]].replace('\n',' ')) + "," + 
                    repr(f[f_cmatrix[1][i]].replace('\n',' ')) + "," + 
                    repr(str(np.around(cmatrix[f_cmatrix[0][i], f_cmatrix[1][i]], 2))) + ")"])
    
    python_dict = [{'Number of feature(s)':str(nof),
                    'Number of instance(s)':str(noi),
                    'Discrete feature(s)':str(df),
                    'Continuous feature(s)':str(cf),
                    'Unknown feature(s)':str(un),
                    'Missing values':str(mv),
                    'Selected feature':f[descriptive_feature_id-1],
                    'Selected feature values':[str(e).replace('\n','') for e in b0],
                    'Mean':m,
                    'Median':med,
                    'Std':std,
                    'Skewness':sk,
                    'Kurtosis':kurt,
                    'Features':f,
                    'Values':ranges,
                    'Type':t,
                    'Type_2':t2,
                    'Missing values (per feature)':[str(e) for e in mvf],
                    'State':s,
                    'Outlier detection method':outlier_detection_methods[outlier_detection_method_id-1],
                    'Outliers':o,
                    'Compatibility issues':inco,
                    'Features with > 50% missing values':bf_names,
                    'Imputation method':imputation_methods[imputation_method_id-1],
#                    'Features with detected outliers':str(outliers_ind).replace('\n',' '),          
#                    'Number of outliers per feature with detected outliers':str(outliers_num).replace('\n',' '),
#                    'Position of outliers per feature with detected outliers':str(outliers_pos),
                    'Highly correlated pair(s) of features':[str(x).replace('\"', '"') for e in output for x in e],
                    'Meta-attribute(s)':metas_features,
                    }]
    
    d = create_wr_io('curation/json_results.txt', python_dict)
    return d


def data_annotation(data_org):
    c = np.size(data_org,1)
    r = np.size(data_org,0)
    var_type_final = []
    features = []
    ranges = []
    means = []
    var_type_final_2 = []
    incompatibilities = []
    incomp_pos = []

    for j in range(c):
        features.append(data_org.domain.attributes[j].name)
        if(data_org.domain.attributes[j].is_discrete == True):
            y = data_org.domain.attributes[j].values
#            var_type_final_2.append('categorical')
            var_type = []
            y_total = np.zeros(len(y))
            for i in range(len(y)):
                try:
                    y_mod = float(y[i])
                    [y_mod, flag] = formatNumber_v2(y_mod)
                    y_total[i] = y_mod
                    if(flag == 1):
                        var_type.append('int')
                    elif(flag == 0):
                        var_type.append('float')                       
                except:
                    y_mod = y[i]
                    if(isinstance(y_mod,str) == True):
                        var_type.append('string')
                    else:
                        var_type.append('unknown')
                    
            matches_str = [1 for x in var_type if x=='string']
            matches_int = [1 for x in var_type if x=='int']
            matches_float = [1 for x in var_type if x=='float']
            
            if(len(matches_str) == len(var_type)):
                var_type_final.append('string')
                var_type_final_2.append('categorical')
                ranges.append(y)
                incompatibilities.append('no')
                incomp_pos.append('-')
                means.append('-')
            elif(len(matches_int) == len(var_type)):
                if('year' in features[j])|('Date' in features[j])|('date' in features[j])|('year' in features[j])|('yr' in features[j])|('Dates' in features[j])|('Year' in features[j])|('YEAR' in features[j])&('Age' not in features[j]):
                    var_type_final.append('date')
                    var_type_final_2.append('numeric')
                    a = str(formatNumber(np.nanmin(y_total)))+','+str(formatNumber(np.nanmax(y_total)))
                    ranges.append(a.split(','))
                    incompatibilities.append('no')
                    means.append(str(formatNumber(np.around(np.nanmedian(y_total)))))
                else:
                    var_type_final.append('int')
                    a = str(formatNumber(np.nanmin(y_total)))+','+str(formatNumber(np.nanmax(y_total)))
                    ranges.append(a.split(','))
                    incompatibilities.append('no')
                    
                    #NEW condition
                    if(len(str(formatNumber(np.nanmax(y_total)))) == 1)&(len(str(formatNumber(np.nanmin(y_total)))) == 1):
                        var_type_final_2.append('categorical')
                        means.append(str(formatNumber(np.around(np.nanmedian(y_total)))))
                    else:
                        var_type_final_2.append('numeric')
                        means.append(str(formatNumber(np.around(np.nanmean(y_total),2))))
                incomp_pos.append('-')
            elif((len(matches_int)+len(matches_float) == len(var_type))&(len(matches_float)!=0)&(len(matches_int)!=0)):
                if('year' in features[j])|('Date' in features[j])|('date' in features[j])|('year' in features[j])|('yr' in features[j])|('Dates' in features[j])|('Year' in features[j])|('YEAR' in features[j])&('Age' not in features[j]):
                    var_type_final.append('date')
                    var_type_final_2.append('numeric')
                    a = str(formatNumber(np.nanmin(y_total)))+','+str(formatNumber(np.nanmax(y_total)))
                    ranges.append(a.split(','))
                    incompatibilities.append('no')
                    means.append(str(formatNumber(np.around(np.nanmedian(y_total)))))
                else:
                    var_type_final.append('float') 
                    var_type_final_2.append('numeric')
                    a = str(formatNumber(np.nanmin(y_total)))+','+str(formatNumber(np.nanmax(y_total)))
                    ranges.append(a.split(','))
                    incompatibilities.append('no') 
                    means.append(str(formatNumber(np.around(np.nanmean(y_total),2))))
                incomp_pos.append('-')             
            elif(len(matches_float) == len(var_type)):
                if('year' in features[j])|('Date' in features[j])|('date' in features[j])|('year' in features[j])|('yr' in features[j])|('Dates' in features[j])|('Year' in features[j])|('YEAR' in features[j])&('Age' not in features[j]):
                    var_type_final.append('date') 
                    var_type_final_2.append('numeric')
                    a = str(formatNumber(np.nanmin(y_total)))+','+str(formatNumber(np.nanmax(y_total)))
                    ranges.append(a.split(',')) 
                    incompatibilities.append('no') 
                    means.append(str(formatNumber(np.around(np.nanmedian(y_total)))))
                else:
                    var_type_final.append('float') 
                    var_type_final_2.append('numeric')
                    a = str(formatNumber(np.nanmin(y_total)))+','+str(formatNumber(np.nanmax(y_total)))
                    ranges.append(a.split(',')) 
                    incompatibilities.append('no') 
                    means.append(str(formatNumber(np.around(np.nanmean(y_total),2))))
                incomp_pos.append('-')
            else:
                var_type_final.append('unknown') 
                var_type_final_2.append('unknown')
                ranges.append(y) 
                means.append('-')
                p = [i for i,x in enumerate(var_type) if ((x=='string')|(x=='unknown'))]
                incomp_names = [y[k] for k in p]
                incomp_names_ind = []
                for m in range(len(incomp_names)):
                    incomp_names_ind.append([i for i,x in enumerate(data_org[:,j]) if x.list[0]==incomp_names[m]])
                incomp_names_ind = list(itertools.chain.from_iterable(incomp_names_ind))
                incomp_pos.append(incomp_names_ind)
                incompatibilities.append("yes, unknown type of data")
        else:
#            var_type_final_2.append('numeric')
            if('year' in features[j])|('Date' in features[j])|('date' in features[j])|('year' in features[j])|('yr' in features[j])|('Dates' in features[j])|('Year' in features[j])|('YEAR' in features[j])&('Age' not in features[j]):
                var_type_final.append('date') 
                means.append(str(formatNumber(np.around(np.nanmedian(data_org.X[:,j])))))
                a = str(formatNumber(np.nanmin(data_org.X[:,j])))+','+str(formatNumber(np.nanmax(data_org.X[:,j])))
                ranges.append(a.split(',')) 
                incompatibilities.append('no') 
                incomp_pos.append('-')
                var_type_final_2.append('numeric') #new
            else:
                g = np.array(list(data_org.X[:,j]))
                g = g[np.isnan(g) == False]
                if(np.all(g%1==0)):
                    var_type_final.append('int')
                    a = str(formatNumber(np.nanmin(data_org.X[:,j])))+','+str(formatNumber(np.nanmax(data_org.X[:,j])))
                    ranges.append(a.split(',')) 
                    incompatibilities.append('no') 
                    incomp_pos.append('-')
                    
                    #new condition
                    if(len(str(formatNumber(np.nanmax(data_org.X[:,j])))) == 1)&(len(str(formatNumber(np.nanmin(data_org.X[:,j])))) == 1):
                        var_type_final_2.append('categorical') 
                        means.append(str(formatNumber(np.around(np.nanmedian(data_org.X[:,j])))))
                    else:
                        var_type_final_2.append('numeric') 
                        means.append(str(formatNumber(np.around(np.nanmean(data_org.X[:,j]),2))))
                else:
                    var_type_final.append('float')
                    a = str(formatNumber(np.nanmin(data_org.X[:,j])))+','+str(formatNumber(np.nanmax(data_org.X[:,j])))
                    ranges.append(a.split(',')) 
                    incompatibilities.append('no') 
                    incomp_pos.append('-')
                    
                    #new condition
                    if(len(str(formatNumber(np.nanmax(data_org.X[:,j])))) == 1)&(len(str(formatNumber(np.nanmin(data_org.X[:,j])))) == 1):
                        var_type_final_2.append('categorical') 
                        means.append(str(formatNumber(np.around(np.nanmedian(data_org.X[:,j])))))
                    else:
                        var_type_final_2.append('numeric') 
                        means.append(str(formatNumber(np.around(np.nanmean(data_org.X[:,j]),2))))
                        
    return [r, c, var_type_final, var_type_final_2, ranges, incompatibilities, incomp_pos, means] 


def handle_duplicates(path_old):
    path_new = path_old.split('/')[1].split('.')[0]+'.xlsx'
    workbook = Workbook(path_new)
    worksheet = workbook.add_worksheet()
    with open(path_old, 'rt', encoding='utf8') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)
    workbook.close()
    
    wb = open_workbook(path_new)
    sheet = wb.sheet_by_index(0)

    #rename duplicated fetures
    features = sheet.row_values(0)
    counts = Counter(features) # so we have: {'name':3, 'state':1, 'city':1, 'zip':2}
    for s,num in counts.items():
       if num > 1: # ignore strings that only appear once
           for suffix in range(1, num + 1): # suffix starts at 1 and increases by 1 each time
               features[features.index(s)] = s + str(suffix) # replace each appearance of s
          
    rb = copy(wb)
    s = rb.get_sheet(0)
    col = 0
    for item in features:
        s.write(0, col, item)
        col += 1
    rb.save(path_new)

    return path_new


def dataframe_to_orange(df):
    variables = []
    for column in df.columns:
        if df[column].dtype.kind in 'biufc':  # If the column is of type binary, integer, unsigned int, float, complex
            variables.append(Orange.data.ContinuousVariable(column))
        else:
            unique_values = df[column].astype(str).unique().tolist()
            variables.append(Orange.data.DiscreteVariable(column, values=unique_values))

    # Construct domain
    domain = Orange.data.Domain(variables)

    # Convert DataFrame to numpy array and handle missing data
    data = df.astype(str).values  # Convert all data to string
    table = Orange.data.Table.from_list(domain, data.tolist())
    return table


def save_to_excel(df, filepath):
    df.to_excel(filepath, index=False)


def load_with_openpyxl(filepath):
    wb = load_workbook(filepath)
    return wb


def cleanup_temp_files(filepath):
    os.remove(filepath)


def data_curator(file_stream, file_extension, imputation_method_id, outlier_detection_method_id, sim_method_id):
    print('Imputation method id:', str(imputation_method_id))
    print('Outlier detection method id:', str(outlier_detection_method_id))
    print('Similarity detection method id:', str(sim_method_id))
    
    print("Loading data from memory...")
    start = timeit.default_timer()

    if file_extension == 'csv':
        df = pd.read_csv(StringIO(file_stream.decode('utf-8')))
    elif file_extension == 'json':
        data = json.loads(file_stream.decode('utf-8'))
        df = pd.DataFrame(data['data'], columns=data['features'])
    elif file_extension in ['xlsx', 'xlsm', 'xltx', 'xltm']:
        wb = load_workbook(BytesIO(file_stream))
        sheet = wb.active
        data = sheet.values
        columns = next(data)[0:]
        df = pd.DataFrame(data, columns=columns)
    else:
        raise ValueError("Unsupported file format.")
    
    features = df.columns
    ncols = df.shape[1]
    nrows = df.shape[0]
    
    data_org = dataframe_to_orange(df)

    print("Constructing the paths for the output...")
    # Generate a filename based on the uploaded file name (fallback to 'dataset' if not available)
    fname = request.files['file'].filename if 'file' in request.files else 'dataset'
    fname_base = fname.rsplit('.', 1)[0]  # Remove file extension

    # Construct output filenames
    fname_c = f"{fname_base}_curated_dataset.xlsx"
    fname_c_v2 = f"{fname_base}_curated_dataset_clean.xlsx"
    fname_e = f"{fname_base}_evaluation_report.xlsx"
    fname_sim_corr = f"{fname_base}_similarity_report_corr.xlsx"
    fname_simlex = f"{fname_base}_similarity_report_lex.xlsx"
    fname_json = f"{fname_base}_results.json"

    path_f_c = 'results/'+fname_c
    path_f_c_v2 = 'results/'+fname_c_v2
    path_f_e = 'results/'+fname_e
    path_f_sim_corr = 'results/'+fname_sim_corr
    path_f_sim_lex = 'results/'+fname_simlex
    path_f_json = 'results/'+fname_json

    print("1. Path for the curated dataset:", fname_c)
    print("2. Path for the clean curated dataset:", fname_c_v2)
    print("3. Path for the quality evaluation report:", fname_e)
    print("4. Path for the similarity report with highly correlated features:", fname_sim_corr)
    print("5. Path for the similarity report with lexically similar features:", fname_simlex)
    print("6. Path for the results json:", fname_json)
    print("")

    if(imputation_method_id is None):
        imputation_method_id = 0
        
    if(outlier_detection_method_id is None):
        outlier_detection_method_id = 1
    
    if(sim_method_id == 5):
        sim_method_id = None

    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec')
    print()

    print("Annotating data...")
    start = timeit.default_timer()
    [r, c, var_type_final, var_type_final_2, ranges, incompatibilities, incomp_pos, means] = data_annotation(data_org)
    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec')
    print()
    
    print("Assessing the quality of the data...")
    start = timeit.default_timer()
    th = len(data_org)/3 #THIS PARAMETER CONTROLS THE PERCENTAGE OF MISSING VALUES
    print("Threshold for missing values:", str(th))
    [features_missing_values, bad_features, bad_features_ind, fair_features, fair_features_ind, 
     good_features, good_features_ind, features_state, a, a_total] = QualityAssessment_S1(data_org, c, th)
    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec') 
    print()
    
    print("Checking for outliers...")
    start = timeit.default_timer()    
    [outliers_ind, y_score, outliers_pos] = outliers_detection(data_org, c, var_type_final, outlier_detection_method_id, 
                                                               bad_features_ind, fair_features_ind, good_features_ind, 
                                                               features_missing_values)
    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec')
    print()

    print("Handling meta-attribute(s)...")
    start = timeit.default_timer()
    [features_total, metas_features, pos_metas, y_total, var_type_metas, var_type_metas_2, 
     features_state_metas, incompatibilities_metas, features_missing_values_metas, bad_features_metas, 
     bad_features_ind_metas, fair_features_metas, fair_features_ind_metas, good_features_metas, 
     good_features_ind_metas, a_total_metas, outliers_ind_metas, y_score_metas, outliers_pos_metas, 
     ranges_metas, incomp_pos_metas, means_metas] = metas_handler(data_org, features, outlier_detection_method_id, df)
    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec'); print()

    D_c_json = None
    D_j_json = None
    
    if(sim_method_id is not None):
        print("Applying similarity detection...")
        start = timeit.default_timer()
        np.seterr(divide='ignore', invalid='ignore')
        [cmatrix, f_cmatrix, _, _, _, _, _, _, _, _, _, D_c, D_j] = similarity_detection(wb, int(sim_method_id))
        stop = timeit.default_timer()
        print('Time: ', np.around(stop - start, 3), 'sec')
        print()
        
        print("Creating the similarity detection report...")
        start = timeit.default_timer()
        
        D_c.to_excel(path_f_sim_corr, index=False)
        D_j.to_excel(path_f_sim_lex, index=False)

        D_c_json = D_c.to_dict('list')
        D_j_json = D_j.to_dict('list')

        stop = timeit.default_timer()
        print('Time: ', np.around(stop - start, 3), 'sec')
        print()
        
    print("Creating the evaluation report...")
    start = timeit.default_timer()
    write_evaluation_report(data_org, r, c, features_total, metas_features, pos_metas, ranges, var_type_final, var_type_final_2, 
                            var_type_metas, var_type_metas_2, features_state_metas, incompatibilities_metas, features_missing_values_metas, 
                            bad_features_metas, bad_features_ind_metas,fair_features_metas, fair_features_ind_metas, 
                            good_features_metas,good_features_ind_metas, a_total_metas, outliers_ind_metas, y_score_metas, outliers_pos, 
                            outliers_pos_metas, ranges_metas, features_missing_values, features_state, outliers_ind, incompatibilities, a_total, path_f_e)
    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec') 
    print()
    
    print("Creating the curated dataset...")
    start = timeit.default_timer()
    write_curated_dataset(data_org, wb, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                          var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                          var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, path_f_c, var_type_final_2)     
    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec')
    print()

    print("Creating the clean curated dataset...")
    start = timeit.default_timer()
    write_curated_dataset_v2(data_org, wb, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                             var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                             var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, path_f_c_v2, var_type_final_2)  
    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec')
    print()

    print("Creating a holistic JSON file...")
    start = timeit.default_timer()
    print("JSONifying the evaluation report...")
    evaluation_report_dict = write_evaluation_report_dict(data_org, r, c, features_total, metas_features, 
                                                        pos_metas, ranges, var_type_final, var_type_final_2, 
                                                        var_type_metas, var_type_metas_2, features_state_metas, 
                                                        incompatibilities_metas, features_missing_values_metas, 
                                                        bad_features_metas, bad_features_ind_metas,fair_features_metas, 
                                                        fair_features_ind_metas, good_features_metas, good_features_ind_metas, 
                                                        a_total_metas, outliers_ind_metas, y_score_metas, outliers_pos_metas, 
                                                        ranges_metas, features_missing_values, features_state, outliers_ind, 
                                                        incompatibilities, a_total, means, means_metas, 
                                                        outlier_detection_method_id, imputation_method_id)
    
    print("JSONifying the curated dataset...")
    curated_dataset_dict = write_curated_dataset_dict(data_org, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                                                    var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                                                    var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, var_type_final_2, df, ncols, nrows)
    

    print("JSONifying the clean curated dataset...")
    curated_dataset_clean_dict = write_curated_dataset_v2_dict(data_org, pos_metas, features_total, features_state, metas_features, imputation_method_id, 
                                                             var_type_final, outliers_pos, r, c, features_state_metas, outliers_pos_metas, var_type_metas, 
                                                             var_type_metas_2, y_total, incomp_pos, incomp_pos_metas, var_type_final_2, df, ncols-len(evaluation_report_dict[0]["Bad features"]), nrows)

    print("Merging everything...")
    timestamp = datetime.now()
    my_datetime_str = timestamp.isoformat()
    final_json = []
    final_json.append({"timestamp": my_datetime_str})
    
    for i in evaluation_report_dict:
        final_json.append(i)
        
    for i in curated_dataset_dict:
        final_json.append(i)

    for i in curated_dataset_clean_dict:
        final_json.append(i)
    
    final_json.append(D_c_json)
    final_json.append(D_j_json)

    save_file = open(path_f_json, "w")
    json.dump(final_json, save_file, indent=6)  
    save_file.close()  

    stop = timeit.default_timer()
    print('Time: ', np.around(stop - start, 3), 'sec')
    print()
    print("Done!")


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/main', methods=['GET', 'POST'])
def index():
    success = False
    message = ""
    if request.method == 'POST':
        if 'file' not in request.files:
            message = 'No file attached in request'
            return render_template('index.html', success=success, message=message)
        
        file = request.files['file']
        if file.filename == '':
            message = 'No file selected'
            return render_template('index.html', success=success, message=message)
        
        if file and allowed_file(file.filename):
            try:
                imputation_method_id = request.form.get("imputation_method_id")
                outlier_detection_method_id = request.form.get("outlier_detection_method_id")
                sim_method_id = request.form.get("sim_method_id")

                # Read file into memory
                file_extension = file.filename.rsplit('.', 1)[1].lower()
                file_stream = file.stream.read()

                data_curator(file_stream, 
                             file_extension,
                             int(imputation_method_id), 
                             int(outlier_detection_method_id), 
                             int(sim_method_id))
                
                message = "Successful execution! Navigate to the folder 'results/'"
                success = True
            except Exception as e:
                message = str(e)

    return render_template('index.html', success=success, message=message)


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)
    